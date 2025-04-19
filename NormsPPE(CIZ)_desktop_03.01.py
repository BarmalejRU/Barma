import sqlite3
from tkinter import Tk, Label, Entry, Button, Text, END, Listbox, Menu, Toplevel, Frame, StringVar, messagebox, BooleanVar, Checkbutton, Canvas
from tkinter import ttk
import os
import sys
import logging
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import re
import os
import logging
from datetime import datetime

# Функция для записи основных событий
def log_event(event_type, message):
    if event_type == "info":
        logging.info(message)
    elif event_type == "warning":
        logging.warning(message)
    elif event_type == "error":
        logging.error(message)

# Определение пути к папке для логов
def get_logs_directory():
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    else:
        application_path = os.path.dirname(os.path.abspath(__file__))
    logs_dir = os.path.join(application_path, "Logs")
    os.makedirs(logs_dir, exist_ok=True)  # Создаём папку "Logs", если её нет
    return logs_dir

# Настройка логирования
LOGS_DIR = get_logs_directory()
LOG_FILE_PATH = os.path.join(LOGS_DIR, "program_history.log")

logging.basicConfig(
    filename=LOG_FILE_PATH,          # Файл для записи логов
    level=logging.INFO,             # Логируем все события от INFO и выше
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Глобальные переменные для сбора статистики
session_stats = {
    "start_time": None,
    "end_time": None,
    "cards_created": 0,
    "errors_count": 0,
    "search_attempts": 0,
    "invalid_inputs": 0
}

# Функция для записи основных событий
def log_event(event_type, message):
    if event_type == "info":
        logging.info(message)
    elif event_type == "warning":
        logging.warning(message)
    elif event_type == "error":
        logging.error(message)

# Функция для обновления статистики
def update_session_stats(stats_key):
    session_stats[stats_key] += 1

# Функция для завершения сеанса
def end_session():
    session_stats["end_time"] = datetime.now()
    duration = session_stats["end_time"] - session_stats["start_time"]
    log_event("info", f"Программа завершена. Статистика сеанса: "
               f"Карточек создано: {session_stats['cards_created']}, "
               f"Ошибок: {session_stats['errors_count']}, "
               f"Поисков: {session_stats['search_attempts']}, "
               f"Некорректных вводов: {session_stats['invalid_inputs']}, "
               f"Продолжительность: {duration}")

# Запись начала сеанса
session_stats["start_time"] = datetime.now()
log_event("info", "Программа запущена.")

# Глобальный обработчик исключений
def handle_uncaught_exception(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    error_message = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    logging.error(f"Uncaught exception: {error_message}")
    print(f"❌ Произошла ошибка: {error_message}")

sys.excepthook = handle_uncaught_exception

# Настройка логирования
logging.basicConfig(filename='program_log.txt', level=logging.ERROR,
                    format='%(asctime)s:%(levelname)s:%(message)s')

# Определение пути к базе данных
def get_database_path():
    return r"D:\pythonProject1\NormaCIZ\norma_ciz.db"

DATABASE_FILE = get_database_path()

# Очистка имени файла для безопасного сохранения
def sanitize_filename(filename):
    sanitized = re.sub(r'[\\/*?:"<>|\r]', '', filename)
    sanitized = sanitized.strip().replace(' ', '_')
    max_length = 50
    if len(sanitized) > max_length:
        sanitized = sanitized[:max_length]
    return sanitized

# Словари сопоставления полей и ячеек
left_cells = {
    "Фамилия": "A8",
    "Имя": "A9",
    "Отчество": "A10",
    "Табельный номер": "A11",
    "Структурное подразделение": "A12",
    "Профессия": "A13",
    "Дата поступления": "A14",
    "Дата изменения профессии": "A15"
}
right_cells = {
    "Пол": "C8",
    "Рост": "C9",
    "Размер одежды": "C11",
    "Размер обуви": "C12",
    "Размер гол. убора": "C13",
    "Размер СИЗОД": "C14",
    "Размер СИЗ рук": "C15",
    "Примечание": "C16"
}

# Функция для переключения фокуса на следующий виджет
def focus_next_entry(event):
    """
    Переключает фокус на следующий виджет при нажатии клавиши "Enter".
    """
    event.widget.tk_focusNext().focus()
    return "break"

# Подключение к базе данных
def connect_to_database(db_name):
    if not os.path.exists(db_name):
        logging.error(f"База данных '{db_name}' не найдена.")
        print(f"❌ База данных '{db_name}' не найдена.")
        return None
    try:
        conn = sqlite3.connect(db_name)
        return conn
    except Exception as e:
        logging.error(f"Ошибка подключения к базе данных: {e}")
        print(f"Ошибка подключения к базе данных: {e}")
        return None

# Подключение к базе данных
def connect_to_database(db_name):
    if not os.path.exists(db_name):
        logging.error(f"База данных '{db_name}' не найдена.")
        print(f"❌ База данных '{db_name}' не найдена.")
        return None
    try:
        conn = sqlite3.connect(db_name)
        return conn
    except Exception as e:
        logging.error(f"Ошибка подключения к базе данных: {e}")
        print(f"Ошибка подключения к базе данных: {e}")
        return None

# Поиск подсказок
def fetch_suggestions(event):
    user_input = entry.get().strip()
    if not user_input:
        suggestions_listbox.delete(0, END)
        return
    conn = connect_to_database(DATABASE_FILE)
    if not conn:
        return
    cursor = conn.cursor()
    try:
        cursor.execute('SELECT DISTINCT professiya FROM normy_vydachi WHERE professiya COLLATE NOCASE LIKE ? LIMIT 10',
                       (f'%{user_input}%',))
        results = cursor.fetchall()
        suggestions_listbox.delete(0, END)
        for result in results:
            suggestions_listbox.insert(END, result[0])
    except Exception as e:
        logging.error(f"Ошибка при поиске подсказок: {e}")
        print(f"Ошибка при поиске подсказок: {e}")
    finally:
        if conn:
            conn.close()

# Выбор подсказки
def select_suggestion(event):
    selected_index = suggestions_listbox.curselection()
    if not selected_index:
        return
    selected_text = suggestions_listbox.get(selected_index[0])
    entry.delete(0, END)
    entry.insert(0, selected_text)
    suggestions_listbox.delete(0, END)

# Расчет норм СИЗ
def calculate_siz():
    work_description = entry.get().strip()
    result_text.delete(1.0, END)
    if not work_description:
        result_text.insert(END, "⚠️ Введите название должности.\n")
        return
    conn = connect_to_database(DATABASE_FILE)
    if not conn:
        result_text.insert(END, "❌ База данных недоступна.\n")
        return
    cursor = conn.cursor()
    try:
        cursor.execute('''
        WITH RankedData AS (
            SELECT 
                id,
                professiya,
                tip_sredstva,
                naimenovanie,
                edinitsa_izmereniya,
                kolichestvo,
                MAX(CASE WHEN professiya != '' THEN id END) OVER (ORDER BY id) AS last_profession_id
            FROM normy_vydachi
        )
        SELECT 
            tip_sredstva,
            naimenovanie,
            edinitsa_izmereniya,
            kolichestvo
        FROM RankedData
        WHERE last_profession_id = (
            SELECT MIN(id) FROM normy_vydachi WHERE professiya = ?
        )
        AND (tip_sredstva != '' OR naimenovanie != '' OR edinitsa_izmereniya != '' OR kolichestvo != '')
        ORDER BY id ASC
        ''', (work_description,))
        results = cursor.fetchall()
        if results:
            result_text.insert(END, f"  Нормы выдачи СИЗ для: '{work_description}':\n")
            result_text.tag_configure("red", foreground="red")
            for tip_sredstva, naimenovanie, edinitsa_izmereniya, kolichestvo in results:
                result_text.insert(END, f"  - [{tip_sredstva}] {naimenovanie}: ")
                result_text.insert(END, f"{kolichestvo}\n", "red")
            global last_results, last_profession
            last_results = [(ts, nm, ei, kl) for ts, nm, ei, kl in results if any((ts, nm, ei, kl))]
            last_profession = work_description
        else:
            result_text.insert(END, f"⚠️ Для должности '{work_description}' нормы не найдены.\n")
    except Exception as e:
        result_text.insert(END, f"❌ Произошла ошибка: {e}\n")
        logging.error(f"Ошибка при расчете норм: {e}")
    finally:
        if conn:
            conn.close()

# Создание карточки сотрудника
def create_siz_card_with_employee_data(data, results):
    try:
        base_dir = os.path.dirname(DATABASE_FILE)
        cards_dir = os.path.join(base_dir, "Cards")
        os.makedirs(cards_dir, exist_ok=True)
        template_path = os.path.join(cards_dir, "Карточка СИЗ.xlsx")
        if not os.path.exists(template_path):
            error_msg = ("Файл шаблона 'Карточка СИЗ.xlsx' не найден!\n"
                         f"Полный путь: {template_path}\n"
                         "Что проверить:\n"
                         "1. Файл должен быть в папке 'Cards'\n"
                         "2. Название должно быть точным (включая регистр букв)\n"
                         "3. Файл не должен быть открыт в Excel")
            logging.error(error_msg)  # Логируем ошибку
            print(f"❌ Ошибка: {error_msg}")  # Выводим в консоль
            messagebox.showerror("Ошибка - файл не найден", error_msg)
            return None
        workbook = load_workbook(template_path)
        sheet = workbook.active

        # Левый столбец
        left_formats = {
            "Фамилия": "Фамилия:",
            "Имя": "Имя:",
            "Отчество": "Отчество (при наличии):",
            "Табельный номер": "Табельный номер:",
            "Структурное подразделение": "Структурное подразделение:",
            "Профессия": "Профессия (должность):",
            "Дата поступления": "Дата поступления на работу:",
            "Дата изменения профессии": "Дата изменения профессии (должности) или перевода в другое структурное подразделение:"
        }
        for field, cell_address in left_cells.items():
            cell = sheet[cell_address]
            prefix = left_formats.get(field, "")
            value = data.get(field, "")
            cell.value = f"{prefix}\t\t{value}" if value else prefix
            cell.font = Font(bold=True)
            if field == "Дата изменения профессии":
                cell.alignment = Alignment(wrap_text=True)

        # Правый столбец
        right_formats = {
            "Пол": "Пол:",
            "Рост": "Рост:",
            "Размер одежды": "Размер:\nодежды -",
            "Размер обуви": "обуви -",
            "Размер гол. убора": "гол.убора -",
            "Размер СИЗОД": "СИЗОД:",
            "Размер СИЗ рук": "СИЗ рук:",
            "Примечание": "Примечание:"
        }
        for field, cell_address in right_cells.items():
            cell = sheet[cell_address]
            prefix = right_formats.get(field, "")
            value = data.get(field, "")
            if field in ["Размер одежды", "Размер обуви", "Размер гол. убора"]:
                cell.value = f"{prefix} {value}" if value else prefix
            else:
                cell.value = f"{prefix}\t\t{value}" if value else prefix
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)

        # Обработка новых полей
        if data.get("responsible_check") and data.get("responsible_person"):
            sheet['C128'] = data["responsible_person"]
            sheet.merge_cells('C128:D128')  # Используем merge_cells вместо merge_range
        if data.get("card_check") and data.get("card_number"):
            sheet['A3'] = f"ЛИЧНАЯ КАРТОЧКА № {data['card_number']}"
            sheet.merge_cells('A3:D3')  # Используем merge_cells вместо merge_range

        # Фильтруем пустые записи
        filtered_results = [row for row in results if any(row)]

        # Заполнение таблицы СИЗ
        # Заполнение таблицы СИЗ
        start_row = 20
        for idx, row in enumerate(filtered_results):
            if len(row) != 4:
                error_msg = f"Некорректные данные в строке {idx + 1}: {row}"
                logging.error(error_msg)
                print(f"❌ Ошибка: {error_msg}")
                messagebox.showerror("Ошибка", error_msg)
                return None

            protection_type, item_name, unit, quantity = row
            current_row = start_row + idx

            # Записываем наименование СИЗ
            sheet[f'A{current_row}'] = item_name or "Нет данных"
            sheet[f'A{current_row}'].alignment = Alignment(wrap_text=True, horizontal='left')

            # Записываем единицу измерения
            sheet[f'C{current_row}'] = unit or "Нет данных"
            sheet[f'C{current_row}'].alignment = Alignment(wrap_text=True, horizontal='left')

            # Записываем количество на период
            try:
                # Пытаемся преобразовать quantity в число
                numeric_quantity = float(quantity) if quantity else None
                sheet[f'D{current_row}'] = numeric_quantity
            except ValueError:
                # Если преобразование не удалось, записываем как текст
                sheet[f'D{current_row}'] = str(quantity) if quantity else "Нет данных"
                sheet[f'D{current_row}'].number_format = "@"  # Явно указываем текстовый формат

            # Добавляем примечание
            sheet[f"B{current_row}"] = "П.2.1.1 Приложение N 2 "
            sheet[f'B{current_row}'].alignment = Alignment(wrap_text=True, horizontal='left')

        # Генерация имени файла
        filename = sanitize_filename(f"{data['Фамилия']}_{data['Имя']}.xlsx")
        save_path = os.path.join(cards_dir, filename)
        counter = 1
        while os.path.exists(save_path):
            new_filename = sanitize_filename(f"{data['Фамилия']}_{data['Имя']}_{counter}.xlsx")
            save_path = os.path.join(cards_dir, new_filename)
            counter += 1

        # Сохранение файла
        workbook.save(save_path)
        messagebox.showinfo("Успех", f"Карточка сохранена:\n{save_path}")
        return save_path

    except Exception as e:
        error_msg = f"Ошибка при создании карточки:\n{str(e)}\n{traceback.format_exc()}"
        logging.error(error_msg)  # Логируем ошибку
        print(f"❌ Ошибка: {error_msg}")  # Выводим в консоль
        messagebox.showerror("Ошибка", error_msg)
        return None

# Создание окна ввода данных сотрудника
def create_employee_window():
    global employee_data, entries, employee_window
    if 'employee_window' in globals() and employee_window is not None and employee_window.winfo_exists():
        employee_window.lift()  # Поднимаем существующее окно на передний план
        employee_window.focus_force()
        return
    employee_data = {}
    entries = {}
    employee_window = Toplevel(root)
    employee_window.title("Заполнение данных сотрудника")
    employee_window.geometry("900x700")
    employee_window.protocol("WM_DELETE_WINDOW", on_employee_window_close)
    x = root.winfo_x() + root.winfo_width()
    y = root.winfo_y()
    employee_window.geometry(f"+{x}+{y}")
    main_frame = Frame(employee_window)
    main_frame.pack(fill='both', expand=True, padx=10, pady=10)
    canvas = Canvas(main_frame)
    scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = Frame(canvas)
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    left_frame = Frame(scrollable_frame)
    left_frame.pack(side="left", padx=10)
    right_frame = Frame(scrollable_frame)
    right_frame.pack(side="right", padx=10)

    def validate_and_save():
        data = {
            "Фамилия": entries.get("surname", "").get().strip(),
            "Имя": entries.get("name", "").get().strip(),
            "Отчество": entries.get("patronymic", "").get().strip(),
            "Табельный номер": entries.get("tab_number", "").get().strip(),
            "Структурное подразделение": entries.get("department", "").get().strip(),
            "Профессия": last_profession,
            "Дата поступления": entries.get("hire_date", "").get().strip(),
            "Дата изменения профессии": entries.get("change_date", "").get().strip(),
            "Пол": entries.get("gender", "").get().strip().upper(),
            "Рост": entries.get("height", "").get().strip(),
            "Размер одежды": entries.get("clothes_size", "").get().strip(),
            "Размер обуви": entries.get("shoes_size", "").get().strip(),
            "Размер гол. убора": entries.get("helmet_size", "").get().strip(),
            "Размер СИЗОД": entries.get("sizod_size", "").get().strip(),
            "Размер СИЗ рук": entries.get("gloves_size", "").get().strip(),
            "Примечание": entries.get("note", "").get().strip(),
            "responsible_person": entries.get("responsible_person", "").get().strip(),
            "responsible_check": entries["responsible_check"].var.get(),
            "card_number": entries.get("card_number", "").get().strip(),
            "card_check": entries["card_check"].var.get()
        }
        errors = []
        if not data["Фамилия"] or not data["Фамилия"].isalpha():
            errors.append("❌ Фамилия должна содержать только буквы.")
            entries["surname"].config(bg="red")
        else:
            entries["surname"].config(bg="white")
        if not data["Имя"] or not data["Имя"].isalpha():
            errors.append("❌ Имя должно содержать только буквы.")
            entries["name"].config(bg="red")
        else:
            entries["name"].config(bg="white")
        if not data["Табельный номер"].isdigit():
            errors.append("❌ Табельный номер должен быть числом.")
            entries["tab_number"].config(bg="red")
        else:
            entries["tab_number"].config(bg="white")
        def is_valid_date(date_str):
            try:
                day, month, year = map(int, date_str.split('.'))
                return 1 <= day <= 31 and 1 <= month <= 12
            except:
                return False
        if not is_valid_date(data["Дата поступления"]):
            errors.append("❌ Некорректная дата поступления (формат: DD.MM.YYYY).")
            entries["hire_date"].config(bg="red")
        else:
            entries["hire_date"].config(bg="white")
        if data["Дата изменения профессии"] and not is_valid_date(data["Дата изменения профессии"]):
            errors.append("❌ Некорректная дата изменения профессии (формат: DD.MM.YYYY).")
            entries["change_date"].config(bg="red")
        else:
            entries["change_date"].config(bg="white")
        if data["Пол"] not in ["М", "Ж"]:
            errors.append("❌ Пол должен быть 'М' или 'Ж'.")
            style.map("TCombobox", fieldbackground=[("readonly", "red")])
        else:
            style.map("TCombobox", fieldbackground=[("readonly", "white")])
        numeric_fields = {
            "Рост": "height",
            "Размер одежды": "clothes_size",
            "Размер обуви": "shoes_size",
            "Размер гол. убора": "helmet_size"
        }
        for field, key in numeric_fields.items():
            value = entries.get(key, "")
            if value and not value.get().strip().isdigit():
                errors.append(f"❌ Поле '{field}' должно быть числом.")
                entries[key].config(bg="red")
            else:
                entries[key].config(bg="white")
        if errors:
            error_message = "\n".join(errors)
            messagebox.showwarning("Ошибка", f"Исправьте следующие ошибки:\n{error_message}")
            return
        missing_fields = [field for field, value in data.items() if
                          field not in ["Фамилия", "Имя", "Табельный номер", "Дата поступления", "Размер СИЗОД",
                                        "Размер СИЗ рук", "responsible_person", "responsible_check", "card_number",
                                        "card_check"] and not value]
        if missing_fields:
            confirm = messagebox.askyesno("Подтверждение",
                                          "Не все поля заполнены. Хотите продолжить?")
            if not confirm:
                return
        file_path = create_siz_card_with_employee_data(data, last_results)
        if file_path:
            result_text.insert(END, f"✅ Карточка успешно создана: {file_path}\n")

    def clear_employee_entries():
        for key, entry_widget in entries.items():
            if isinstance(entry_widget, ttk.Combobox):
                entry_widget.set('')
            elif isinstance(entry_widget, (Entry, ttk.Entry)):
                # Проверяем состояние чекбоксов перед очисткой
                if key == "responsible_person":
                    if not entries["responsible_check"].var.get():  # Очищаем только если чекбокс НЕ активен
                        entry_widget.delete(0, END)
                        entry_widget.config(bg='white')
                elif key == "card_number":
                    if not entries["card_check"].var.get():  # Очищаем только если чекбокс НЕ активен
                        entry_widget.delete(0, END)
                        entry_widget.config(bg='white')
                else:
                    entry_widget.delete(0, END)
                    entry_widget.config(bg='white')
            elif isinstance(entry_widget, Checkbutton):
                # Не сбрасываем состояние чекбокса
                continue

    # Левая колонка
    Label(left_frame, text="Фамилия:").pack(anchor="w")
    entries["surname"] = Entry(left_frame, width=30)
    entries["surname"].pack(pady=5)
    entries["surname"].bind("<Return>", focus_next_entry)
    Label(left_frame, text="Имя:").pack(anchor="w")
    entries["name"] = Entry(left_frame, width=30)
    entries["name"].pack(pady=5)
    entries["name"].bind("<Return>", focus_next_entry)
    Label(left_frame, text="Отчество (при наличии):").pack(anchor="w")
    entries["patronymic"] = Entry(left_frame, width=30)
    entries["patronymic"].pack(pady=5)
    entries["patronymic"].bind("<Return>", focus_next_entry)
    Label(left_frame, text="Табельный номер:").pack(anchor="w")
    entries["tab_number"] = Entry(left_frame, width=30)
    entries["tab_number"].pack(pady=5)
    entries["tab_number"].bind("<Return>", focus_next_entry)
    Label(left_frame, text="Структурное подразделение:").pack(anchor="w")
    entries["department"] = Entry(left_frame, width=30)
    entries["department"].pack(pady=5)
    entries["department"].bind("<Return>", focus_next_entry)
    Label(left_frame, text="Дата поступления на работу (DD.MM.YYYY):").pack(anchor="w")
    entries["hire_date"] = Entry(left_frame, width=30)
    entries["hire_date"].pack(pady=5)
    entries["hire_date"].bind("<Return>", focus_next_entry)
    Label(left_frame, text="Дата изменения профессии (DD.MM.YYYY):").pack(anchor="w")
    entries["change_date"] = Entry(left_frame, width=30)
    entries["change_date"].pack(pady=5)
    entries["change_date"].bind("<Return>", focus_next_entry)

    # Правая колонка
    Label(right_frame, text="Пол:").pack(anchor="w")
    gender_var = StringVar()
    entries["gender"] = ttk.Combobox(right_frame, textvariable=gender_var, values=["М", "Ж"], width=28)
    entries["gender"].pack(pady=5)
    Label(right_frame, text="Рост:").pack(anchor="w")
    entries["height"] = Entry(right_frame, width=30)
    entries["height"].pack(pady=5)
    entries["height"].bind("<Return>", focus_next_entry)
    Label(right_frame, text="Размер одежды:").pack(anchor="w")
    entries["clothes_size"] = Entry(right_frame, width=30)
    entries["clothes_size"].pack(pady=5)
    entries["clothes_size"].bind("<Return>", focus_next_entry)
    Label(right_frame, text="Размер обуви:").pack(anchor="w")
    entries["shoes_size"] = Entry(right_frame, width=30)
    entries["shoes_size"].pack(pady=5)
    entries["shoes_size"].bind("<Return>", focus_next_entry)
    Label(right_frame, text="Размер гол. убора:").pack(anchor="w")
    entries["helmet_size"] = Entry(right_frame, width=30)
    entries["helmet_size"].pack(pady=5)
    entries["helmet_size"].bind("<Return>", focus_next_entry)
    Label(right_frame, text="Размер СИЗОД:").pack(anchor="w")
    entries["sizod_size"] = Entry(right_frame, width=30)
    entries["sizod_size"].pack(pady=5)
    entries["sizod_size"].bind("<Return>", focus_next_entry)
    Label(right_frame, text="Размер СИЗ рук:").pack(anchor="w")
    entries["gloves_size"] = Entry(right_frame, width=30)
    entries["gloves_size"].pack(pady=5)
    entries["gloves_size"].bind("<Return>", focus_next_entry)
    Label(right_frame, text="Примечание:").pack(anchor="w")
    entries["note"] = Entry(right_frame, width=30)
    entries["note"].pack(pady=5)
    entries["note"].bind("<Return>", focus_next_entry)

    # Новые поля с чекбоксами
    separator = ttk.Separator(scrollable_frame, orient='horizontal')
    separator.pack(fill='x', pady=10)

    # Поле 1: Ответственное лицо
    resp_frame = Frame(scrollable_frame)
    resp_frame.pack(fill='x', padx=10, pady=5)
    Label(resp_frame, text="Ответственное лицо за ведение карточек учета выдачи СИЗ:").pack(anchor='w')
    entries["responsible_person"] = Entry(resp_frame, width=50)
    entries["responsible_person"].pack(side='left', fill='x', expand=True)
    entries["responsible_check"] = Checkbutton(resp_frame)
    entries["responsible_check"].var = BooleanVar()
    entries["responsible_check"]["variable"] = entries["responsible_check"].var
    entries["responsible_check"].pack(side='right', padx=5)

    # Поле 2: Личная карточка №
    card_frame = Frame(scrollable_frame)
    card_frame.pack(fill='x', padx=10, pady=5)
    Label(card_frame, text="ЛИЧНАЯ КАРТОЧКА №").pack(anchor='w')
    entries["card_number"] = Entry(card_frame, width=20)
    entries["card_number"].pack(side='left')
    entries["card_check"] = Checkbutton(card_frame)
    entries["card_check"].var = BooleanVar()
    entries["card_check"]["variable"] = entries["card_check"].var
    entries["card_check"].pack(side='left', padx=5)

    # Кнопки внизу
    button_frame = Frame(scrollable_frame)
    button_frame.pack(fill='x', pady=10)
    Button(button_frame, text="Создать карточку", command=validate_and_save).pack(side='left', padx=5)
    Button(button_frame, text="Очистить данные", command=clear_employee_entries).pack(side='left', padx=5)

def on_employee_window_close():
    global employee_window
    if 'employee_window' in globals() and employee_window is not None and employee_window.winfo_exists():
        employee_window.destroy()
    employee_window = None

def export_to_excel():
    if not last_results or not last_profession:
        result_text.insert(END, "⚠️ Нет данных для экспорта. Сначала выполните расчет.\n")
        return
    create_employee_window()

def clear_fields():
    entry.delete(0, END)
    result_text.delete(1.0, END)

    if 'employee_window' in globals() and employee_window is not None and employee_window.winfo_exists():
        logging.info("Очистка полей в окне заполнения данных сотрудника...")

        for key, widget in entries.items():
            logging.debug(f"Обработка поля: {key}")

            # === Очищаем Combobox
            if isinstance(widget, ttk.Combobox):
                widget.set("")

            # === Очищаем Entry и ttk.Entry, кроме исключений
            elif isinstance(widget, (Entry, ttk.Entry)):
                if key == "responsible_person":
                    var = entries.get("responsible_check_var")
                    if var and var.get():
                        logging.debug("Пропущено поле 'responsible_person' — чекбокс активен.")
                        continue
                if key == "card_number":
                    var = entries.get("card_check_var")
                    if var and var.get():
                        logging.debug("Пропущено поле 'card_number' — чекбокс активен.")
                        continue

                widget.delete(0, END)
                widget.config(bg='white')

        logging.info("Очистка завершена.")



def show_context_menu(event):
    context_menu.post(event.x_root, event.y_root)

# Создание графического интерфейса
root = Tk()
root.title("Расчет норм выдачи СИЗ")
root.geometry("600x500")

# Глобальные переменные
last_results = []
last_profession = ""
employee_window = None
entries = {}

# Стили
style = ttk.Style()
style.configure("TCombobox", fieldbackground="white")

# Виджеты основного окна
Label(root, text="Введите название должности:").pack()
entry = Entry(root, width=50)
entry.pack()
suggestions_listbox = Listbox(root, height=5, width=50)
suggestions_listbox.pack()
Button(root, text="Рассчитать", command=calculate_siz).pack(pady=5)
Button(root, text="Напечатать карточку", command=export_to_excel).pack(pady=5)
Button(root, text="Очистить", command=clear_fields).pack(pady=5)

# Текстовое поле для вывода результатов
result_text = Text(root, height=15, width=70)
result_text.pack()

# Контекстное меню
context_menu = Menu(result_text, tearoff=0)
context_menu.add_command(label="Копировать", command=lambda: result_text.event_generate("<<Copy>>"))
result_text.bind("<Button-3>", show_context_menu)

# Привязки событий
entry.bind('<KeyRelease>', fetch_suggestions)
suggestions_listbox.bind('<<ListboxSelect>>', select_suggestion)
result_text.bind("<Control-c>", lambda event: result_text.event_generate("<<Copy>>"))
result_text.bind("<Control-a>", lambda event: result_text.tag_add("sel", "1.0", "end"))

# Запуск программы
root.mainloop()