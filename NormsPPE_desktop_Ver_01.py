import sqlite3
from tkinter import Tk, Label, Entry, Button, Text, END, Listbox, Menu, Toplevel, Frame, StringVar, messagebox, \
    BooleanVar, Checkbutton, Canvas
from tkinter import ttk
import os
import sys
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import re
import logging
from datetime import datetime

# Словари сопоставления полей и ячеек
left_cells = {
    "Фамилия": "A8",
    "Имя": "A9",
    "Отчество": "A10",
    "Табельный номер": "A11",
    "Структурное подразделение": "A12",
    "Профессия": "A13",
    "Дата поступления": "A14",
    "Дата изменения профессии": "A15"
}
right_cells = {
    "Пол": "C8",
    "Рост": "C9",
    "Размер одежды": "C11",
    "Размер обуви": "C12",
    "Размер гол. убора": "C13",
    "Размер СИЗОД": "C14",
    "Размер СИЗ рук": "C15",
    "Примечание": "C16"
}


# Настройка логирования
def setup_logging():
    def get_logs_directory():
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        else:
            application_path = os.path.dirname(os.path.abspath(__file__))
        logs_dir = os.path.join(application_path, "Logs")
        os.makedirs(logs_dir, exist_ok=True)
        return logs_dir

    LOGS_DIR = get_logs_directory()
    LOG_FILE_PATH = os.path.join(LOGS_DIR, "program_history.log")
    logging.basicConfig(
        filename=LOG_FILE_PATH,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )


# Глобальные переменные
class AppState:
    def __init__(self):
        self.last_results = []
        self.last_profession = ""
        self.employee_window = None
        self.entries = {}
        self.session_stats = {
            "start_time": datetime.now(),
            "end_time": None,
            "cards_created": 0,
            "errors_count": 0,
            "search_attempts": 0,
            "invalid_inputs": 0
        }


# Обработка ошибок
def setup_exception_handler():
    def handle_uncaught_exception(exc_type, exc_value, exc_traceback):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
            return
        error_message = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
        logging.error(f"Uncaught exception: {error_message}")
        print(f"❌ Произошла ошибка: {error_message}")

    sys.excepthook = handle_uncaught_exception


# Работа с базой данных
class DatabaseManager:
    @staticmethod
    def get_database_path():
        return r"D:\pythonProject1\NormaCIZ\norma_ciz.db"

    @staticmethod
    def connect_to_database(db_name):
        if not os.path.exists(db_name):
            logging.error(f"База данных '{db_name}' не найдена.")
            print(f"❌ База данных '{db_name}' не найдена.")
            return None
        try:
            conn = sqlite3.connect(db_name)
            return conn
        except Exception as e:
            logging.error(f"Ошибка подключения к базе данных: {e}")
            print(f"Ошибка подключения к базе данных: {e}")
            return None


# Вспомогательные функции
def sanitize_filename(filename):
    sanitized = re.sub(r'[\\/*?:"<>|\r]', '', filename)
    sanitized = sanitized.strip().replace(' ', '_')
    return sanitized[:50]


def focus_next_entry(event):
    event.widget.tk_focusNext().focus()
    return "break"


class MainApplication:
    def __init__(self, root):
        self.root = root
        self.state = AppState()
        self.setup_ui()
        self.setup_styles()

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.configure("TCombobox", fieldbackground="white")

    def setup_ui(self):
        self.root.title("Расчет норм выдачи СИЗ")
        self.root.geometry("600x500")
        # Основные элементы интерфейса
        Label(self.root, text="Введите название должности:").pack()
        self.entry = Entry(self.root, width=50)
        self.entry.pack()
        self.suggestions_listbox = Listbox(self.root, height=5, width=50)
        self.suggestions_listbox.pack()
        Button(self.root, text="Рассчитать", command=self.calculate_siz).pack(pady=5)
        Button(self.root, text="Напечатать карточку", command=self.export_to_excel).pack(pady=5)
        Button(self.root, text="Очистить", command=self.clear_fields).pack(pady=5)
        # Текстовое поле для результатов
        self.result_text = Text(self.root, height=15, width=70)
        self.result_text.pack()
        # Контекстное меню
        self.setup_context_menu()
        # Привязки событий
        self.entry.bind('<KeyRelease>', self.fetch_suggestions)
        self.suggestions_listbox.bind('<<ListboxSelect>>', self.select_suggestion)
        self.result_text.bind("<Control-c>", lambda e: self.result_text.event_generate("<<Copy>>"))
        self.result_text.bind("<Control-a>", lambda e: self.result_text.tag_add("sel", "1.0", "end"))

    def setup_context_menu(self):
        context_menu = Menu(self.result_text, tearoff=0)
        context_menu.add_command(label="Копировать", command=lambda: self.result_text.event_generate("<<Copy>>"))
        self.result_text.bind("<Button-3>", lambda e: context_menu.post(e.x_root, e.y_root))

    # Основные функции приложения
    def fetch_suggestions(self, event):
        user_input = self.entry.get().strip()
        if not user_input:
            self.suggestions_listbox.delete(0, END)
            return
        conn = DatabaseManager.connect_to_database(DatabaseManager.get_database_path())
        if not conn:
            return
        try:
            cursor = conn.cursor()
            cursor.execute(
                'SELECT DISTINCT professiya FROM normy_vydachi WHERE professiya COLLATE NOCASE LIKE ? LIMIT 10',
                (f'%{user_input}%',))
            self.suggestions_listbox.delete(0, END)
            for result in cursor.fetchall():
                self.suggestions_listbox.insert(END, result[0])
        except Exception as e:
            logging.error(f"Ошибка при поиске подсказок: {e}")
            print(f"Ошибка при поиске подсказок: {e}")
        finally:
            conn.close()

    def select_suggestion(self, event):
        if not (selected_index := self.suggestions_listbox.curselection()):
            return
        selected_text = self.suggestions_listbox.get(selected_index[0])
        self.entry.delete(0, END)
        self.entry.insert(0, selected_text)
        self.suggestions_listbox.delete(0, END)

    def calculate_siz(self):
        work_description = self.entry.get().strip()
        self.result_text.delete(1.0, END)
        if not work_description:
            self.result_text.insert(END, "⚠️ Введите название должности.\n")
            return
        conn = DatabaseManager.connect_to_database(DatabaseManager.get_database_path())
        if not conn:
            self.result_text.insert(END, "❌ База данных недоступна.\n")
            return
        try:
            cursor = conn.cursor()
            cursor.execute('''
            WITH RankedData AS (
                SELECT 
                    id,
                    professiya,
                    tip_sredstva,
                    naimenovanie,
                    edinitsa_izmereniya,
                    kolichestvo,
                    MAX(CASE WHEN professiya != '' THEN id END) OVER (ORDER BY id) AS last_profession_id
                FROM normy_vydachi
            )
            SELECT 
                tip_sredstva,
                naimenovanie,
                edinitsa_izmereniya,
                kolichestvo
            FROM RankedData
            WHERE last_profession_id = (
                SELECT MIN(id) FROM normy_vydachi WHERE professiya = ?
            )
            AND (tip_sredstva != '' OR naimenovanie != '' OR edinitsa_izmereniya != '' OR kolichestvo != '')
            ORDER BY id ASC
            ''', (work_description,))
            results = cursor.fetchall()
            if results:
                self.result_text.insert(END, f"  Нормы выдачи СИЗ для: '{work_description}':\n")
                self.result_text.tag_configure("red", foreground="red")
                for tip_sredstva, naimenovanie, edinitsa_izmereniya, kolichestvo in results:
                    self.result_text.insert(END, f"  - [{tip_sredstva}] {naimenovanie}: ")
                    self.result_text.insert(END, f"{kolichestvo}\n", "red")
                self.state.last_results = [(ts, nm, ei, kl) for ts, nm, ei, kl in results if any((ts, nm, ei, kl))]
                self.state.last_profession = work_description
            else:
                self.result_text.insert(END, f"⚠️ Для должности '{work_description}' нормы не найдены.\n")
        except Exception as e:
            self.result_text.insert(END, f"❌ Произошла ошибка: {e}\n")
            logging.error(f"Ошибка при расчете норм: {e}")
        finally:
            conn.close()

    def create_employee_window(self):
        if self.state.employee_window and self.state.employee_window.winfo_exists():
            self.state.employee_window.lift()
            self.state.employee_window.focus_force()
            return

        # Создаем новое окно
        self.state.employee_window = Toplevel(self.root)
        self.state.employee_window.title("Заполнение данных сотрудника")
        self.state.employee_window.geometry("900x500")
        self.state.employee_window.protocol("WM_DELETE_WINDOW", self.on_employee_window_close)

        # Получаем размеры основного окна и экрана
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Вычисляем позицию второго окна
        window_width = 900
        window_height = 500
        offset = 50

        x = root_x + root_width + offset
        y = root_y

        if x + window_width > screen_width:
            x = screen_width - window_width - offset
        if y + window_height > screen_height:
            y = screen_height - window_height - offset

        self.state.employee_window.geometry(f"+{x}+{y}")

        # Основной контейнер с прокруткой
        main_frame = Frame(self.state.employee_window)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        canvas = Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Создаем колонки
        left_frame = Frame(scrollable_frame)
        left_frame.pack(side="left", padx=5)
        right_frame = Frame(scrollable_frame)
        right_frame.pack(side="right", padx=5)

        # Инициализируем entries
        self.state.entries = {}
        self.create_left_column(left_frame)
        self.create_right_column(right_frame)
        self.create_additional_fields(scrollable_frame)
        self.create_action_buttons(scrollable_frame)

        # Устанавливаем фокус на первое поле
        self.state.entries["surname"].focus_set()

    def create_left_column(self, parent):
        fields = [
            ("Фамилия:", "surname"),
            ("Имя:", "name"),
            ("Отчество (при наличии):", "patronymic"),
            ("Табельный номер:", "tab_number"),
            ("Структурное подразделение:", "department"),
            ("Дата поступления на работу (DD.MM.YYYY):", "hire_date"),
            ("Дата изменения профессии (DD.MM.YYYY):", "change_date")
        ]

        def adjust_entry_width(event):
            widget = event.widget
            text_length = len(widget.get())
            new_width = max(30, text_length + 5)
            widget.config(width=new_width)

        for label_text, field_name in fields:
            field_container = Frame(parent)
            field_container.pack(fill='x', pady=5, anchor='w')

            Label(field_container, text=label_text).pack(anchor='w')

            input_row = Frame(field_container)
            input_row.pack(fill='x')

            entry = Entry(input_row, width=30)
            entry.pack(side='left', fill='x', expand=True)
            entry.bind("<Return>", self.custom_focus_next_entry)
            self.state.entries[field_name] = entry

            if field_name == "department":
                self.state.entries["department_check_var"] = BooleanVar()
                self.state.entries["department_check"] = Checkbutton(
                    input_row,
                    variable=self.state.entries["department_check_var"]
                )
                self.state.entries["department_check"].pack(side='left', padx=5)

            entry.bind("<KeyRelease>", adjust_entry_width)

    def create_right_column(self, parent):
        Label(parent, text="").pack(anchor="w", pady=5)

        Label(parent, text="Пол:").pack(anchor="w")
        gender_var = StringVar()
        self.state.entries["gender"] = ttk.Combobox(parent, textvariable=gender_var, values=["М", "Ж"], width=28)
        self.state.entries["gender"].pack(pady=5)

        fields = [
            ("Рост:", "height"),
            ("Размер одежды:", "clothes_size"),
            ("Размер обуви:", "shoes_size"),
            ("Размер гол. убора:", "helmet_size"),
            ("Размер СИЗОД:", "sizod_size"),
            ("Размер СИЗ рук:", "gloves_size"),
            ("Примечание:", "note")
        ]
        for label_text, field_name in fields:
            Label(parent, text=label_text).pack(anchor="w")
            entry = Entry(parent, width=30)
            entry.pack(pady=5)
            entry.bind("<Return>", self.custom_focus_next_entry)
            self.state.entries[field_name] = entry

        def adjust_entry_width(event):
            widget = event.widget
            text_length = len(widget.get())
            new_width = max(30, text_length + 5)
            widget.config(width=new_width)

        for field_name in self.state.entries:
            widget = self.state.entries[field_name]
            if isinstance(widget, (Entry, ttk.Entry)):
                widget.bind("<KeyRelease>", adjust_entry_width)

    def create_additional_fields(self, parent):
        ttk.Separator(parent, orient='horizontal').pack(fill='x', pady=10)

        resp_frame = Frame(parent)
        resp_frame.pack(fill='x', padx=10, pady=5)

        Label(resp_frame, text="Ответственное лицо за ведение карточек учета выдачи СИЗ:").pack(anchor='w')

        self.state.entries["responsible_person"] = Entry(resp_frame, width=50)
        self.state.entries["responsible_person"].pack(side='left', fill='x', expand=True)
        self.state.entries["responsible_person"].bind("<Return>", self.custom_focus_next_entry)

        self.state.entries["responsible_check_var"] = BooleanVar()
        self.state.entries["responsible_check"] = Checkbutton(
            resp_frame,
            variable=self.state.entries["responsible_check_var"]
        )
        self.state.entries["responsible_check"].pack(side='right', padx=5)

        card_frame = Frame(parent)
        card_frame.pack(fill='x', padx=10, pady=5)

        Label(card_frame, text="ЛИЧНАЯ КАРТОЧКА №").pack(anchor='w')

        self.state.entries["card_number"] = Entry(card_frame, width=20)
        self.state.entries["card_number"].pack(side='left')
        self.state.entries["card_number"].bind("<Return>", self.custom_focus_next_entry)

        self.state.entries["card_check_var"] = BooleanVar()
        self.state.entries["card_check"] = Checkbutton(
            card_frame,
            variable=self.state.entries["card_check_var"]
        )
        self.state.entries["card_check"].pack(side='left', padx=5)

    def create_action_buttons(self, parent):
        button_frame = Frame(parent)
        button_frame.pack(fill='x', pady=10)
        Button(button_frame, text="Создать карточку", command=self.validate_and_save).pack(side='left', padx=5)
        Button(button_frame, text="Очистить данные", command=self.clear_employee_entries).pack(side='left', padx=5)
        Button(button_frame, text="Закрыть", command=self.on_employee_window_close).pack(side='left', padx=5)

    def custom_focus_next_entry(self, event):
        current_widget = event.widget
        next_widget = current_widget.tk_focusNext()

        while isinstance(next_widget, Checkbutton):
            next_widget = next_widget.tk_focusNext()

        if not next_widget:
            first_widget = self.state.entries["surname"]
            first_widget.focus_set()
            return "break"

        next_widget.focus_set()
        return "break"

    def clear_employee_entries(self):
        try:
            protected_fields = {
                "department": "department_check_var",
                "responsible_person": "responsible_check_var",
                "card_number": "card_check_var"
            }

            for key, widget in self.state.entries.items():
                if isinstance(widget, ttk.Combobox):
                    widget.set('')
                elif isinstance(widget, (Entry, ttk.Entry)):
                    if key in protected_fields and self.state.entries[protected_fields[key]].get():
                        continue
                    widget.delete(0, END)
                    widget.config(bg='white')
                elif isinstance(widget, Checkbutton):
                    if key not in {"department_check", "responsible_check", "card_check"}:
                        if hasattr(widget, 'var'):
                            widget.var.set(False)
                        elif hasattr(widget, 'variable'):
                            widget.variable.set(False)

            self.state.entries["surname"].focus_set()

        except Exception as e:
            logging.error(f"Ошибка при очистке данных: {e}")
            print(f"❌ Ошибка при очистке данных: {e}")
            messagebox.showerror("Ошибка", f"Произошла ошибка при очистке данных: {e}")

    def validate_and_save(self):
        data = {
            "Фамилия": self.state.entries["surname"].get().strip(),
            "Имя": self.state.entries["name"].get().strip(),
            "Отчество": self.state.entries["patronymic"].get().strip(),
            "Табельный номер": self.state.entries["tab_number"].get().strip(),
            "Структурное подразделение": self.state.entries["department"].get().strip(),
            "Профессия": self.state.last_profession,
            "Дата поступления": self.state.entries["hire_date"].get().strip(),
            "Дата изменения профессии": self.state.entries["change_date"].get().strip(),
            "Пол": self.state.entries["gender"].get().strip().upper(),
            "Рост": self.state.entries["height"].get().strip(),
            "Размер одежды": self.state.entries["clothes_size"].get().strip(),
            "Размер обуви": self.state.entries["shoes_size"].get().strip(),
            "Размер гол. убора": self.state.entries["helmet_size"].get().strip(),
            "Размер СИЗОД": self.state.entries["sizod_size"].get().strip(),
            "Размер СИЗ рук": self.state.entries["gloves_size"].get().strip(),
            "Примечание": self.state.entries["note"].get().strip(),
            "responsible_person": self.state.entries["responsible_person"].get().strip(),
            "responsible_check": self.state.entries["responsible_check_var"].get(),
            "card_number": self.state.entries["card_number"].get().strip(),
            "card_check": self.state.entries["card_check_var"].get(),
            "department_check": self.state.entries["department_check_var"].get()
        }

        errors = self.validate_employee_data(data)
        if errors:
            messagebox.showwarning("Ошибка", f"Исправьте следующие ошибки:\n{errors}")
            return

        if not self.confirm_incomplete_data(data):
            return

        file_path = self.create_siz_card_with_employee_data(data)
        if file_path:
            self.result_text.insert(END, f"✅ Карточка успешно создана: {file_path}\n")
            self.state.session_stats["cards_created"] += 1

    def validate_employee_data(self, data):
        errors = []
        if not data["Фамилия"] or not data["Фамилия"].isalpha():
            errors.append("❌ Фамилия должна содержать только буквы.")
            self.state.entries["surname"].config(bg="red")
        else:
            self.state.entries["surname"].config(bg="white")

        if not data["Имя"] or not data["Имя"].isalpha():
            errors.append("❌ Имя должно содержать только буквы.")
            self.state.entries["name"].config(bg="red")
        else:
            self.state.entries["name"].config(bg="white")

        if not data["Табельный номер"].isdigit():
            errors.append("❌ Табельный номер должен быть числом.")
            self.state.entries["tab_number"].config(bg="red")
        else:
            self.state.entries["tab_number"].config(bg="white")

        if not self.is_valid_date(data["Дата поступления"]):
            errors.append("❌ Некорректная дата поступления (формат: DD.MM.YYYY).")
            self.state.entries["hire_date"].config(bg="red")
        else:
            self.state.entries["hire_date"].config(bg="white")

        if data["Дата изменения профессии"] and not self.is_valid_date(data["Дата изменения профессии"]):
            errors.append("❌ Некорректная дата изменения профессии (формат: DD.MM.YYYY).")
            self.state.entries["change_date"].config(bg="red")
        else:
            self.state.entries["change_date"].config(bg="white")

        if data["Пол"] not in ["М", "Ж"]:
            errors.append("❌ Пол должен быть 'М' или 'Ж'.")
            self.style.map("TCombobox", fieldbackground=[("readonly", "red")])
        else:
            self.style.map("TCombobox", fieldbackground=[("readonly", "white")])

        numeric_fields = {
            "Рост": "height",
            "Размер одежды": "clothes_size",
            "Размер обуви": "shoes_size",
            "Размер гол. убора": "helmet_size"
        }
        for field, key in numeric_fields.items():
            value = self.state.entries[key].get().strip()
            if value and not value.isdigit():
                errors.append(f"❌ Поле '{field}' должно быть числом.")
                self.state.entries[key].config(bg="red")
            else:
                self.state.entries[key].config(bg="white")
        return "\n".join(errors) if errors else None

    def is_valid_date(self, date_str):
        try:
            day, month, year = map(int, date_str.split('.'))
            return 1 <= day <= 31 and 1 <= month <= 12
        except:
            return False

    def confirm_incomplete_data(self, data):
        missing_fields = [field for field, value in data.items()
                          if field not in ["Фамилия", "Имя", "Табельный номер", "Дата поступления",
                                           "Размер СИЗОД", "Размер СИЗ рук", "responsible_person",
                                           "responsible_check", "card_number", "card_check",
                                           "department_check"] and not value]
        if missing_fields:
            return messagebox.askyesno("Подтверждение",
                                       "Не все поля заполнены. Хотите продолжить?")
        return True

    def create_siz_card_with_employee_data(self, data):
        try:
            base_dir = os.path.dirname(DatabaseManager.get_database_path())
            cards_dir = os.path.join(base_dir, "Cards")
            os.makedirs(cards_dir, exist_ok=True)
            template_path = os.path.join(cards_dir, "Карточка СИЗ.xlsx")

            if not os.path.exists(template_path):
                error_msg = ("Файл шаблона 'Карточка СИЗ.xlsx' не найден!\n"
                             f"Полный путь: {template_path}\n"
                             "Что проверить:\n"
                             "1. Файл должен быть в папке 'Cards'\n"
                             "2. Название должно быть точным (включая регистр букв)\n"
                             "3. Файл не должен быть открыт в Excel")
                logging.error(error_msg)
                print(f"❌ Ошибка: {error_msg}")
                messagebox.showerror("Ошибка - файл не найден", error_msg)
                return None

            workbook = load_workbook(template_path)
            sheet = workbook.active

            # Левый столбец
            left_formats = {
                "Фамилия": "Фамилия:",
                "Имя": "Имя:",
                "Отчество": "Отчество (при наличии):",
                "Табельный номер": "Табельный номер:",
                "Структурное подразделение": "Структурное подразделение:",
                "Профессия": "Профессия (должность):",
                "Дата поступления": "Дата поступления на работу:",
                "Дата изменения профессии": "Дата изменения профессии (должности) или перевода в другое структурное подразделение:"
            }
            for field, cell_address in left_cells.items():
                cell = sheet[cell_address]
                prefix = left_formats.get(field, "")
                value = data.get(field, "")
                cell.value = f"{prefix}\t\t{value}" if value else prefix
                cell.font = Font(bold=True)
                if field == "Дата изменения профессии":
                    cell.alignment = Alignment(wrap_text=True)

            # Правый столбец
            right_formats = {
                "Пол": "Пол:",
                "Рост": "Рост:",
                "Размер одежды": "Размер:\nодежды -",
                "Размер обуви": "обуви -",
                "Размер гол. убора": "гол.убора -",
                "Размер СИЗОД": "СИЗОД:",
                "Размер СИЗ рук": "СИЗ рук:",
                "Примечание": "Примечание:"
            }
            for field, cell_address in right_cells.items():
                cell = sheet[cell_address]
                prefix = right_formats.get(field, "")
                value = data.get(field, "")
                if field in ["Размер одежды", "Размер обуви", "Размер гол. убора"]:
                    cell.value = f"{prefix} {value}" if value else prefix
                else:
                    cell.value = f"{prefix}\t\t{value}" if value else prefix
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True)

            # Обработка новых полей
            if data.get("responsible_check") and data.get("responsible_person"):
                sheet['C129'] = data["responsible_person"]
                sheet.merge_cells('C129:D129')
            if data.get("card_check") and data.get("card_number"):
                sheet['A3'] = f"ЛИЧНАЯ КАРТОЧКА № {data['card_number']}"
                sheet.merge_cells('A3:D3')

            # Заполнение таблицы СИЗ
            filtered_results = [row for row in self.state.last_results if any(row)]
            start_row = 20
            for idx, row in enumerate(filtered_results):
                if len(row) != 4:
                    error_msg = f"Некорректные данные в строке {idx + 1}: {row}"
                    logging.error(error_msg)
                    messagebox.showerror("Ошибка", error_msg)
                    return None

                protection_type, item_name, unit, quantity = row
                current_row = start_row + idx

                sheet[f'A{current_row}'] = item_name or "Нет данных"
                sheet[f'C{current_row}'] = unit or "Нет данных"

                try:
                    numeric_quantity = float(quantity) if quantity else None
                    sheet[f'D{current_row}'] = numeric_quantity
                except ValueError:
                    sheet[f'D{current_row}'] = str(quantity) if quantity else "Нет данных"
                    sheet[f'D{current_row}'].number_format = "@"

                sheet[f"B{current_row}"] = "П.2.1.1 Приложение N 2 "
                sheet[f'B{current_row}'].alignment = Alignment(wrap_text=True, horizontal='left')

            # Сохранение файла
            filename = sanitize_filename(f"{data['Фамилия']}_{data['Имя']}.xlsx")
            save_path = os.path.join(cards_dir, filename)

            counter = 1
            while os.path.exists(save_path):
                new_filename = sanitize_filename(f"{data['Фамилия']}_{data['Имя']}_{counter}.xlsx")
                save_path = os.path.join(cards_dir, new_filename)
                counter += 1

            workbook.save(save_path)
            messagebox.showinfo("Успех", f"Карточка сохранена:\n{save_path}")
            return save_path

        except Exception as e:
            error_msg = f"Ошибка при создании карточки:\n{str(e)}\n{traceback.format_exc()}"
            logging.error(error_msg)
            print(f"❌ Ошибка: {error_msg}")
            messagebox.showerror("Ошибка", error_msg)
            return None

    def on_employee_window_close(self):
        if self.state.employee_window and self.state.employee_window.winfo_exists():
            self.state.employee_window.destroy()
        self.state.employee_window = None

    def export_to_excel(self):
        if not self.state.last_results or not self.state.last_profession:
            self.result_text.insert(END, "⚠️ Нет данных для экспорта. Сначала выполните расчет.\n")
            return
        self.create_employee_window()

    def clear_fields(self):
        self.entry.delete(0, END)
        self.result_text.delete(1.0, END)
        if self.state.employee_window and self.state.employee_window.winfo_exists():
            self.clear_employee_entries()


# Запуск приложения
if __name__ == "__main__":
    setup_logging()
    setup_exception_handler()
    root = Tk()
    app = MainApplication(root)
    root.mainloop()
    app.state.session_stats["end_time"] = datetime.now()
    duration = app.state.session_stats["end_time"] - app.state.session_stats["start_time"]
    logging.info(f"Программа завершена. Статистика сеанса: "
                 f"Время работы: {duration}, "
                 f"Создано карточек: {app.state.session_stats['cards_created']}, "
                 f"Ошибок: {app.state.session_stats['errors_count']}, "
                 f"Поисков: {app.state.session_stats['search_attempts']}, "
                 f"Невалидных вводов: {app.state.session_stats['invalid_inputs']}")